"""Dev-only tool: regenerate asin_group_mapping.csv from ASIN_MAPPING_ALL.xlsx.

Not part of the deployed Cloud Function - GetSalesDepartmentReport.py only reads the
committed CSV output at runtime. Requires openpyxl locally (not in requirements.txt).

Usage:
    python3 build_asin_mapping.py <path-to-ASIN_MAPPING_ALL.xlsx> [output-csv]
"""
import csv
import sys

import openpyxl

DEFAULT_OUTPUT = "asin_group_mapping.csv"


def load_sheet1_rows(ws):
    rows = []
    for ean, asin, sku, per_box, description, group, _extra in ws.iter_rows(min_row=2, values_only=True):
        if not asin or not sku:
            continue
        group_norm = (str(group).strip() if group else "") or "UNGROUPED"
        rows.append({"SKU": str(sku).strip(), "ASIN": str(asin).strip(), "GROUP": group_norm})
    return rows


def load_sheet2_rows(ws):
    rows = []
    for asin, sku in ws.iter_rows(min_row=1, values_only=True):
        if not asin or not sku:
            continue
        rows.append({"SKU": str(sku).strip(), "ASIN": str(asin).strip(), "GROUP": "UNGROUPED"})
    return rows


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 build_asin_mapping.py <path-to-ASIN_MAPPING_ALL.xlsx> [output-csv]", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rows = load_sheet1_rows(wb["Sheet1"])
    if "Sheet2" in wb.sheetnames:
        rows += load_sheet2_rows(wb["Sheet2"])

    seen_skus = set()
    deduped = []
    for row in rows:
        if row["SKU"] in seen_skus:
            print(f"WARNING: duplicate SKU {row['SKU']!r} in mapping, keeping first occurrence", file=sys.stderr)
            continue
        seen_skus.add(row["SKU"])
        deduped.append(row)

    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["SKU", "ASIN", "GROUP"])
        writer.writeheader()
        writer.writerows(deduped)

    print(f"Wrote {len(deduped)} rows to {output_path}")


if __name__ == "__main__":
    main()
